import os
import csv
import tempfile
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import Student, Attendance
from database import db
from sqlalchemy import func

def export_to_excel(class_obj, start_date=None, end_date=None):
    """
    Export attendance data to Excel format with students as rows and dates as columns
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name} Attendance"
    
    # Get all students in the class
    students = Student.query.filter_by(class_id=class_obj.id).order_by(Student.name).all()
    
    # Get all unique dates with attendance records
    query = db.session.query(Attendance.date).filter_by(class_id=class_obj.id)
    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)
    
    attendance_dates = query.distinct().order_by(Attendance.date).all()
    attendance_dates = [date_tuple[0] for date_tuple in attendance_dates]
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create headers
    headers = ["S.No", "Student ID", "Student Name", "Email"]
    headers.extend([date_obj.strftime('%m/%d/%Y') for date_obj in attendance_dates])
    headers.append("Total Present")
    headers.append("Total Absent") 
    headers.append("Total Late")
    headers.append("Attendance %")
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write class information
    ws.cell(row=2, column=1, value="Class:")
    ws.cell(row=2, column=2, value=f"{class_obj.name} - {class_obj.subject}")
    ws.cell(row=3, column=1, value="Teacher:")
    ws.cell(row=3, column=2, value=class_obj.teacher.name)
    ws.cell(row=4, column=1, value="Export Date:")
    ws.cell(row=4, column=2, value=datetime.now().strftime('%m/%d/%Y %I:%M %p'))
    
    # Start student data from row 6
    data_start_row = 6
    
    # Write student data
    for row_idx, student in enumerate(students, data_start_row):
        # Basic student info
        ws.cell(row=row_idx, column=1, value=row_idx - data_start_row + 1)  # Serial number
        ws.cell(row=row_idx, column=2, value=student.student_id)
        ws.cell(row=row_idx, column=3, value=student.name)
        ws.cell(row=row_idx, column=4, value=student.email)
        
        # Get attendance records for this student
        attendance_records = {}
        records = Attendance.query.filter_by(student_id=student.id, class_id=class_obj.id).all()
        for record in records:
            attendance_records[record.date] = record.status
        
        # Statistics
        present_count = 0
        absent_count = 0
        late_count = 0
        
        # Fill attendance data for each date
        for col_idx, date_obj in enumerate(attendance_dates, 5):
            status = attendance_records.get(date_obj, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=status)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            
            # Count attendance and apply color coding
            if status == "Present":
                present_count += 1
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif status == "Absent":
                absent_count += 1
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
            elif status == "Late":
                late_count += 1
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
        
        # Add statistics columns
        total_classes = len(attendance_dates)
        stats_col = len(attendance_dates) + 5
        
        # Total Present
        ws.cell(row=row_idx, column=stats_col, value=present_count).border = border
        # Total Absent
        ws.cell(row=row_idx, column=stats_col + 1, value=absent_count).border = border
        # Total Late
        ws.cell(row=row_idx, column=stats_col + 2, value=late_count).border = border
        # Attendance Percentage
        if total_classes > 0:
            attendance_percentage = round((present_count / total_classes) * 100, 2)
            percentage_cell = ws.cell(row=row_idx, column=stats_col + 3, value=f"{attendance_percentage}%")
            percentage_cell.border = border
            
            # Color code percentage
            if attendance_percentage >= 75:
                percentage_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif attendance_percentage >= 50:
                percentage_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            else:
                percentage_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add legend
    legend_row = len(students) + data_start_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    
    # Present legend
    present_cell = ws.cell(row=legend_row + 1, column=1, value="Present")
    present_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Absent legend
    absent_cell = ws.cell(row=legend_row + 1, column=2, value="Absent")
    absent_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Late legend
    late_cell = ws.cell(row=legend_row + 1, column=3, value="Late")
    late_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

def export_to_csv(class_obj, start_date=None, end_date=None):
    """
    Export attendance data to CSV format with students as rows and dates as columns
    """
    # Get all students in the class
    students = Student.query.filter_by(class_id=class_obj.id).order_by(Student.name).all()
    
    # Get all unique dates with attendance records
    query = db.session.query(Attendance.date).filter_by(class_id=class_obj.id)
    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)
    
    attendance_dates = query.distinct().order_by(Attendance.date).all()
    attendance_dates = [date_tuple[0] for date_tuple in attendance_dates]
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='')
    writer = csv.writer(temp_file)
    
    # Write class information
    writer.writerow(['Class:', f"{class_obj.name} - {class_obj.subject}"])
    writer.writerow(['Teacher:', class_obj.teacher.name])
    writer.writerow(['Export Date:', datetime.now().strftime('%m/%d/%Y %I:%M %p')])
    writer.writerow([])  # Empty row
    
    # Create headers
    headers = ["S.No", "Student ID", "Student Name", "Email"]
    headers.extend([date_obj.strftime('%m/%d/%Y') for date_obj in attendance_dates])
    headers.extend(["Total Present", "Total Absent", "Total Late", "Attendance %"])
    
    writer.writerow(headers)
    
    # Write student data
    for idx, student in enumerate(students, 1):
        # Get attendance records for this student
        attendance_records = {}
        records = Attendance.query.filter_by(student_id=student.id, class_id=class_obj.id).all()
        for record in records:
            attendance_records[record.date] = record.status
        
        # Build row data
        row_data = [idx, student.student_id, student.name, student.email]
        
        # Statistics
        present_count = 0
        absent_count = 0
        late_count = 0
        
        # Add attendance data for each date
        for date_obj in attendance_dates:
            status = attendance_records.get(date_obj, "")
            row_data.append(status)
            
            # Count attendance
            if status == "Present":
                present_count += 1
            elif status == "Absent":
                absent_count += 1
            elif status == "Late":
                late_count += 1
        
        # Add statistics
        total_classes = len(attendance_dates)
        row_data.extend([present_count, absent_count, late_count])
        
        # Add attendance percentage
        if total_classes > 0:
            attendance_percentage = round((present_count / total_classes) * 100, 2)
            row_data.append(f"{attendance_percentage}%")
        else:
            row_data.append("0%")
        
        writer.writerow(row_data)
    
    # Add legend
    writer.writerow([])  # Empty row
    writer.writerow(['Legend:'])
    writer.writerow(['Present = Student was present'])
    writer.writerow(['Absent = Student was absent'])
    writer.writerow(['Late = Student was late'])
    writer.writerow(['Empty = No attendance record for that date'])
    
    temp_file.close()
    return temp_file.name